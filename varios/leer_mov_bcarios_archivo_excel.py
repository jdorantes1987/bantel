import openpyxl
import pandas as pd

ruta = "C:/Users/jdorantes/Documents/Analisis/Estados de Cuenta/Banesco/2023/06. JUN/Banesco 16.06.2023.xlsx"
edo_cta = pd.read_excel(ruta)
# print(edo_cta.info())
wb = openpyxl.load_workbook(ruta)

sheet = wb.worksheets[0]
# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet.cell(row=1, column=1)
# print(cell_obj.value)
# print the total number of rows
# print(sheet.max_row)
# print total number of column
# print(sheet.max_column)
m_row = sheet.max_row
max_col = sheet.max_column
# Loop will print all columns name
# print('Loop will print all columns name')
# for i in range(1, max_col + 1):
#    cell_obj = sheet.cell(row = 1, column = i)
#    print(cell_obj.value)


# Loop will print all values
# of first column
# for i in range(1, m_row + 1):
#    cell_obj = sheet.cell(row = i, column = 2)
#    print(cell_obj.value)

# Write a particular in cell value
# cell_obj = sheet.cell(row = 328, column = 2)
# cell_obj.value = 'Prueba write Python'
# wb.save(ruta)

# Obtiene la primera fila sheet[1]
# first_row = list(sheet.rows)[1]
# Imprime el valor de la columna 1 en la fila 1
# print(first_row[0].value)
# Itera los valores de la fila número 2:
for row in sheet[2]:
    print(row.value)
