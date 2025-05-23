from datetime import date

import comprob as cbte
import numpy as np
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import pandas as pd

from accesos.datos import detalle_comprob, plan_cta
from accesos.files_excel import p_data_comprobantes_manuales as p_cbtes_manual
from administracion_profit.facturas import asiento_conatel
from administracion_profit.obtener_igtf_y_comisiones import (
    contabilizar_comisiones_e_igtf,
)
from varios.utilidades import date_today as hoy
from varios.utilidades import last_date_of_month

wb_load = op.load_workbook(p_cbtes_manual)


def get_data_comprobantes():
    df = detalle_comprob()
    pla_cue = plan_cta()
    # print('Cantidad de valores únicos:', len(df['comp_num'].unique()))
    # id_mto_max = df['monto_d'].idxmax()
    # cta = df.loc[id_mto_max]['co_cue']
    # print(plan_cta.datos_cta(cta)[['co_cue', 'saldo_ini']])
    df2 = pd.merge(df, pla_cue, left_on="co_cue", right_on="co_cue")
    cbt = df2.drop("validador", axis=1)
    # print(cbt.loc[24959:24985][['co_cue', 'des_cue']])
    return cbt


def __insertar_encab_comprob_x_contab():
    sheet = pd.read_excel(p_cbtes_manual, sheet_name="encab")
    sheet_encab_cbtes = wb_load.worksheets[0]
    comprob_x_contab = sheet.replace(np.nan, "")
    comprob_x_contab = comprob_x_contab.query('CONTABILIZAR.str.contains("SI")')
    print(comprob_x_contab.to_string())
    fecha_inset = hoy()
    fecha_mod = hoy()

    for ind in comprob_x_contab.index:
        if (
            str(sheet_encab_cbtes.cell(row=[ind][0] + 2, column=3).value).upper()
            == "SI"
        ):
            col_n_compr = sheet_encab_cbtes.cell(row=[ind][0] + 2, column=1).value
            col_fecha = sheet_encab_cbtes.cell(row=[ind][0] + 2, column=2).value
            col_contabilizar = sheet_encab_cbtes.cell(row=[ind][0] + 2, column=3)
            col_descrip = sheet_encab_cbtes.cell(row=[ind][0] + 2, column=4).value
            #  primero se elimina el comprobante si exite
            cbte.eliminar_comprobante(col_n_compr)
            # Encabezado comprobante
            cbte.new_encab_comprobante(
                col_n_compr, col_descrip, col_fecha, fecha_inset, fecha_mod
            )
            col_contabilizar.value = "NO"


def __insertar_det_comprob_x_contab():
    sh_det = pd.read_excel(p_cbtes_manual, sheet_name="det")
    sh_encab = pd.read_excel(p_cbtes_manual, sheet_name="encab")
    u_sh = pd.merge(
        sh_det, sh_encab, how="left", left_on="ID_CBTE_DET", right_on="ID_CBTE"
    )
    u_sh["CONTABILIZAR"] = u_sh["CONTABILIZAR"].replace(np.nan, " ")
    # Filtra el detalle de los comprobantes a contabilizar
    data = u_sh.query('CONTABILIZAR.str.contains("SI")').copy().reset_index(drop=True)
    # Reemplazar valores nulos de varias columnas
    data[["MTO_DEBE", "MTO_HABER"]] = data[["MTO_DEBE", "MTO_HABER"]].fillna(0)
    data[["ID_CBTE_DET", "ID_CBTE"]] = data[["ID_CBTE_DET", "ID_CBTE"]].astype("Int64")
    data["MTO_HABER"] = abs(data["MTO_HABER"])
    data = data.replace(np.nan, "NULL")  # Quita los valores NaN por espacios
    print(data.to_string())
    for index, row in data.iterrows():
        index += 1
        # Detalle comprobante
        # Se debe validar lo que hay en el campo auxiliar de la tabla detalle comprobante ya es una clave foranea y no permite cadenas vacias sino NULL
        aux = row["CO_AUX"] if row["CO_AUX"] == "NULL" else "'" + row["CO_AUX"] + "'"
        cbte.new_line_comprobante(
            row["ID_CBTE"],
            row["F_EMISION"],
            index,
            row["MTO_DEBE"],
            row["MTO_HABER"],
            row["DET_LINEA"],
            row["CO_CUE"],
            row["DOCREF"],
            hoy(),
            hoy(),
            aux,
        )
    try:
        # Aplicar formato condicional. Regla para resaltar los renglones del comprobante a insertar
        sheet_det_cbtes = wb_load.worksheets[1]
        # Definir el rango de la regla de formato condicional
        rango = "A2:L5000"
        # Definir la fórmula de la regla de formato condicional
        formula = 'VLOOKUP($A2,encab!$A$2:$D$1000,3,FALSE)="SI"'
        # Definir el formato de relleno
        fill = PatternFill(start_color="f8c2c8", end_color="f8c2c8", fill_type="solid")
        # Crear la regla de formato condicional usando la fórmula
        regla = FormulaRule(formula=[formula], fill=fill)
        # Aplicar la regla de formato condicional al rango deseado
        sheet_det_cbtes.conditional_formatting.add(rango, regla)
        wb_load.save(p_cbtes_manual)
        print("Archivo guardado.")
    except Exception as e:
        print("Ha ocurrido un error:", e)
    finally:
        # wb_load.close()
        print("Archivo Cerrado.")


def contabilizar_comprob_manual_file_excel():
    __insertar_encab_comprob_x_contab()
    __insertar_det_comprob_x_contab()


def contabilizar_comprob_conatel(periodo, id_cbte, descrip_encab):
    data = asiento_conatel(periodo)
    fecha = pd.to_datetime(
        last_date_of_month(date(periodo[0], periodo[1], 1))
    ).normalize()  # fecha sin hora, minutos y segundos
    # Encabezado comprobante
    cbte.new_encab_comprobante(id_cbte, descrip_encab, fecha, hoy(), hoy())
    # Detalle comprobante
    for index, row in data.iterrows():
        cbte.new_line_comprobante(
            id_cbte,
            fecha,
            index,
            row["debe"],
            row["haber"],
            row["descrip_det"],
            row["cuenta"],
            "",
            hoy(),
            hoy(),
            "NULL",
        )


if __name__ == "__main__":
    # print(asiento_conatel((2022, 12)))
    # run
    # contabilizar_comisiones_e_igtf(
    #     "2501013", "20250131", "REG. COM. E INTER. BCARIOS ENERO 2025 BANESCO"
    # )
    # run
    contabilizar_comprob_manual_file_excel()

    # # run
    # contabilizar_comprob_conatel(
    #     (2025, 2), "2502018", "PROV. CONATEL Y FONACIT FEB 2025"
    # )
