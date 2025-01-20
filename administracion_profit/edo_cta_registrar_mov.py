from datetime import datetime

from dateutil.relativedelta import relativedelta
from numpy import nan
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pandas import read_excel

from accesos.datos import get_id_movbanco as new_id_mov
from accesos.datos import get_movbanco
from accesos.files_excel import p_data_edo_cta_banesco as p_edo_cta_banesco
from administracion_profit.bancos import new_movbanco as add_new_mov
from administracion_profit.edo_cta import (
    get_mov_bcarios_pdtes_por_identificar,
    get_mov_edo_cta_pdtes_por_regist,
)


def registrar_mov_ban_edo_cta(anio, mes):
    df_edo_cta = read_excel(p_edo_cta_banesco)
    wb_edo_cta = load_workbook(p_edo_cta_banesco)
    sheet = wb_edo_cta.worksheets[0]
    fecha_fin = datetime(anio, mes, 1) + relativedelta(
        day=31
    )  # Obtiene la fecha final del mes

    # m_row = sheet.max_row
    # max_col = sheet.max_column
    df_edo_cta2 = df_edo_cta.replace(nan, "")
    df_edo_cta2["Contabilizar"] = df_edo_cta2["Contabilizar"].str.upper()
    # edo_cta_sin_null = edo_cta_sin_null.query('Comentarios.str.contains("E/S") or Comentarios.str.contains("Botellon")')
    df_edo_cta2 = df_edo_cta2.query('Contabilizar.str.contains("SI")')
    for ind in df_edo_cta2.index:
        if str(sheet.cell(row=[ind][0] + 2, column=10).value).upper() == "SI":
            cell_coll_fecha = sheet.cell(row=[ind][0] + 2, column=1).value
            cell_coll_ref = sheet.cell(row=[ind][0] + 2, column=2).value
            cell_coll_monto = sheet.cell(row=[ind][0] + 2, column=4).value
            cell_coll_descipc = sheet.cell(row=[ind][0] + 2, column=6).value
            cell_coll_cta_ie = sheet.cell(row=[ind][0] + 2, column=9).value
            # cell_coll_idb = round(float(abs(cell_coll_monto) * 0.02), 2) if cell_coll_monto < 0 else 0
            cell_coll_idb = 0.0
            cell_print_result = sheet.cell(row=[ind][0] + 2, column=9)
            cell_contabil = sheet.cell(row=[ind][0] + 2, column=10)
            id_mov_bco = new_id_mov(fecha_fin=fecha_fin)  # Nuevo movimiento bancario
            add_new_mov(
                id_mov_bco,
                cell_coll_descipc,
                cell_coll_cta_ie,
                cell_coll_fecha,
                cell_coll_ref,
                cell_coll_monto,
                cell_coll_idb,
            )
            cell_print_result.value = (
                "CONTABILIZACIÓN-->" + cell_coll_cta_ie + "->" + str(id_mov_bco)
            )
            cell_contabil.value = "NO"
            print(id_mov_bco, df_edo_cta2["Referencia"][ind], [ind][0] + 2, sep="->")
    print("Cantidad de registros a CONTABILIZADOS: ", len(df_edo_cta2.axes[0]))
    try:
        wb_edo_cta.save(p_edo_cta_banesco)
        print("Archivo guardado.")
    except Exception as e:
        print("Ha ocurrido un error:", e)
    finally:
        wb_edo_cta.close()
        wb_edo_cta = None
        print("Archivo Cerrado.")


#  Identifica y establece los movimientos pendientes del estado de cuenta en comprarción con los movimientos bancarios del mes
def establecer_color_amarillo_mov_edo_cta_por_registrar_banesco(fecha_ini):
    #  objecto color de fondo para la celda
    style_amarillo = PatternFill(patternType="solid", fgColor="FFFFB9")
    style_sin_color = PatternFill()
    df = read_excel(p_edo_cta_banesco, dtype={"Referencia": str})
    # Filtra las filas que no tiene fecha
    df_edo_cta = df[df["Fecha"].notnull()].copy()
    # Reemplaza los valores nulos de la columna 'Comentarios' por los valores de la columna 'Descripción'
    df_edo_cta["Comentarios"] = (
        df_edo_cta["Comentarios"].fillna(df_edo_cta["Descripción"]).str[:50]
    )
    wb_edo_cta = load_workbook(p_edo_cta_banesco)
    mov_x_registrar = get_mov_edo_cta_pdtes_por_regist(fecha_ini)
    # Obtiene los movimientos por conciliar
    mov_pdtes = df_edo_cta[df_edo_cta["Referencia"].isin(mov_x_registrar)]
    # Obtiene los movimientos conciliados
    mov_concil = df_edo_cta[~df_edo_cta["Referencia"].isin(mov_x_registrar)]
    # print('Movimientos del edo. cta. Banesco por identificar en libro')
    # print(mov_pdtes.to_string())
    sheet = wb_edo_cta.worksheets[0]
    # Recorre los movimientos conciliados y quita el color de fondo de la celda
    for ind in mov_concil.index:
        cell_col_pendiente = sheet.cell(row=[ind][0] + 2, column=11)
        cell_col_pendiente.value = ""
        estatus = sheet.cell(row=[ind][0] + 2, column=9)
        estatus.fill = style_sin_color
    # Recorre los movimientos por conciliar y aplica color amarillo al fondo de la celda
    for ind in mov_pdtes.index:
        estatus = sheet.cell(row=[ind][0] + 2, column=9)
        estatus.fill = style_amarillo
    try:
        wb_edo_cta.save(p_edo_cta_banesco)
        print("Archivo guardado.")
    except Exception as e:
        print("Ha ocurrido un error:", e)
    finally:
        wb_edo_cta.close()
        print("Proceso terminado!", "Archivo Cerrado.")


#  Este procedimiento hace cruce por REFERENCIA BANCARIA entre el edo cta. banesco y los mov de su cuenta contable en libros
def mov_bcarios_pendientes_por_identif_en_edo_cta_banesco(fecha_ini):
    print("Movimientos bancarios por identificar en estado de cuenta Banesco.")
    ref_por_identif = get_mov_bcarios_pdtes_por_identificar(fecha_ini)
    movimientos = get_movbanco(
        fecha_ini
    )  # Obtiene los movimientos bancarios del mes especificado
    # Obtiene las referencia pendientes por identificar de los movimientos bancarios incluidas en el set 'ref_por_identif'
    mov_pdtes = movimientos[movimientos["doc_num"].isin(ref_por_identif)]
    mov_pdtes_banesco = mov_pdtes[mov_pdtes["cod_cta"] == "0134  "].copy()
    mov_pdtes_banesco["descrip"] = mov_pdtes_banesco["descrip"].str[:40]
    return mov_pdtes_banesco[
        [
            "mov_num",
            "fecha",
            "descrip",
            "co_cta_ingr_egr",
            "doc_num",
            "monto_d",
            "monto_h",
            "cob_pag",
            "co_us_in",
            "fe_us_in",
            "cod_cta",
        ]
    ]
